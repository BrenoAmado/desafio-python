from django.shortcuts import render
from django.shortcuts import redirect
from django.http import JsonResponse
from openpyxl import load_workbook
from .models import Pessoa
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook


def index(request):
    return render(request, 'index.html')


def upload_planilha(request):
    if request.method == 'POST' and request.FILES['file']:
        arquivo = request.FILES['file']
        workbook = load_workbook(arquivo)
        sheet = workbook.active

        # Lista para armazenar nomes e emails das pessoas presentes no excel
        pessoas_no_excel = []

        # Processo de leitura do excel
        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome, email, data_nascimento, ativo = row

            # Verifica se data_nascimento é um objeto datetime, se não, converte a string
            if isinstance(data_nascimento, str):
                data_nascimento = datetime.strptime(data_nascimento, "%Y-%m-%d").date()
            elif isinstance(data_nascimento, datetime):
                data_nascimento = data_nascimento.date()

            idade = (datetime.now().date() - data_nascimento).days // 365

            # Validação de idade e idade perante valor
            if idade < 18:
                ativo = False

            if ativo:
                if idade < 21:
                    valor = 100.00
                elif idade < 60:
                    valor = 150.00
                else:
                    valor = 200.00

                # Verifica se a pessoa já existe com todos os dados iguais
                pessoa_existente = Pessoa.objects.filter(
                    nome=nome,
                    email=email,
                    data_nascimento=data_nascimento,
                    ativo=True,
                    valor=valor
                ).exists()
                
                # Se não existir, cria a nova pessoa
                if not pessoa_existente:
                    Pessoa.objects.create(
                        nome=nome,
                        email=email,
                        data_nascimento=data_nascimento,
                        ativo=True,
                        valor=valor
                    )

                # Adiciona a lista de pessoas atualizada no excel
                pessoas_no_excel.append((nome, email))

            # Exclui pessoas que estão no banco, mas não no Excel
            Pessoa.objects.filter(ativo=True).exclude(
                nome__in=[p[0] for p in pessoas_no_excel],
                email__in=[p[1] for p in pessoas_no_excel]).delete()

            # Recupera todas as pessoas ativas após o processamento
            pessoas_ativas = Pessoa.objects.filter(ativo=True).values(
                'nome', 'email', 'data_nascimento', 'ativo', 'valor')

        # Converte as datas e valores para o formato necessário do JSON
        for pessoa in pessoas_ativas:
            pessoa['data_nascimento'] = pessoa['data_nascimento'].strftime('%Y-%m-%d')
            pessoa['valor'] = float(pessoa['valor'])

        request.session['pessoas_ativas'] = list(pessoas_ativas)
        return redirect('exibir_json')

def exibir_json(request):
    pessoas_ativas = request.session.get('pessoas_ativas', [])
    return JsonResponse(pessoas_ativas, safe=False)


def download_planilha(request):
    pessoas = Pessoa.objects.filter(ativo=True)
    workbook = Workbook()
    sheet = workbook.active

    # Cabeçalho
    sheet.append(["Nome", "Email", "Data de Nascimento", "Ativo", "Valor"])

    # Dados
    for pessoa in pessoas:
        sheet.append([
            pessoa.nome,
            pessoa.email,
            pessoa.data_nascimento,
            pessoa.ativo,
            pessoa.valor
        ])

    # Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Tabela_Atualizada.xlsx'
    workbook.save(response)
    return response
